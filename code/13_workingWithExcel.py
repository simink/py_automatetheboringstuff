### chapter 13 - working with excel spreadsheets

import openpyxl

wb = openpyxl.load_workbook('../docs/example.xlsx')
type(wb)

# get sheetnames
wb.sheetnames

sheet = wb['Sheet3']
sheet
type(sheet)
sheet.title

anotherSheet = wb.active
anotherSheet

## getting cells from sheets
sheet = wb['Sheet1']
sheet['A1'].value

c = sheet['B1']
c.value
print(f'row {c.row}, column {c.column}, value {c.value}, coordinate {c.coordinate}')

## cell method
sheet.cell(row=1, column=2).value

for i in range(1,8,2):  # every other row
    print(i, sheet.cell(row=i, column=2).value)

sheet.max_row
sheet.max_column

## convert between column letters and numbers
from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(1)
get_column_letter(27)
get_column_letter(sheet.max_column)
column_index_from_string('AA')

## getting rows, columns from sheets
tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:
    # print(rowOfCellObjects)
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("---end of row---")

# need to convert attributes to list first to access these values
list(sheet.columns[1])  # get second column cells

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)



### PROJECT: Reading data from spreadsheet

import os
import census2010

census2010.allData['AK']['Anchorage']
anchoragePop = census2010.allData['AK']['Anchorage']['pop']
print('The 2010 population of Anchorage was ' + str(anchoragePop))


### Writing excel documents
import openpyxl
wb = openpyxl.Workbook() # Create a blank workbook.
wb.sheetnames # It starts with one sheet.
sheet = wb.active
sheet.title
sheet.title = 'Spam' # Change title (this is active)
wb.sheetnames
wb.create_sheet()
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet')  # create a new sheet at index 0
wb.sheetnames
del wb['Sheet']
wb.sheetnames
sheet['A1'] = 'Hello, world!'  # writes to spam
wb.save('example_copy.xlsx')


## setting font style of cells
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet['A1'] = 'Hello, world!'

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
# sheet['A1'] = 'Bold Times New Roman'

sheet['B3'] = '24 pt Italic'
fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2

wb.save('styles.xlsx')

### Writing formulas
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
wb.save('writeFormula.xlsx')

### Setting row height column width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

### Merge and unmerge cells
sheet.merge_cells('A1:D3') # Merge all these cells
sheet.unmerge_cells('A1:D3') # Split these cells up

### Freeze panes
sheet.freeze_panes = 'C2'  # row 1 and columns A and B
sheet.freeze_panes = 'A1' or sheet.freeze_panes = None # unfreeze all

### Charts
# Create a Reference object from a rectangular selection of cells.
# Create a Series object by passing in the Reference object.
# Create a Chart object.
# Append the Series object to the Chart object.
# Add the Chart object to the Worksheet object, optionally specifying which cell should be the top-left corner of the chart.

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
    
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')  # placement of top left corner of chart at C5
wb.save('sampleChart.xlsx')








