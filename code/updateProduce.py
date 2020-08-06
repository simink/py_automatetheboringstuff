#! python3
# corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('./docs/produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

### checks prices
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('./docs/updatedProduceSales.xlsx')